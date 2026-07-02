"""
Admin-only handlers: /result (multi-step), /playoff_result, /recalc (with confirm),
/add_match, /admin_log, /setup_api.  All guarded by ADMIN_TELEGRAM_ID.
"""
import csv
import io
import json
from datetime import datetime, timezone, timedelta

from aiogram import Router, F, Bot
from aiogram.filters import Command
from aiogram.types import Message, Document, CallbackQuery

from bot.config import config
from bot.database.db import get_db, fetchone, fetchall
from bot.keyboards import (
    admin_match_list_kb, admin_home_score_kb, admin_away_score_kb,
    admin_playoff_team_kb, admin_playoff_method_kb, admin_confirm_kb,
    recalc_confirm_kb, recalc_match_list_kb,
    playoff_match_list_kb, playoff_pick_team_kb, playoff_pick_method_kb, playoff_confirm_kb,
    api_result_kb,
)
from bot.services.scoring import score_prediction, Prediction, Match
from bot.services.api_client import setup_api, fetch_games, en_to_ru, local_to_msk, ROUND_ORDER

router = Router()
router.message.filter(F.from_user.id == config.admin_telegram_id)
router.callback_query.filter(F.from_user.id == config.admin_telegram_id)

MSK_TZ = timezone(timedelta(hours=3))


# ---------- helpers ----------

def _parse_score(text: str) -> tuple[int, int] | None:
    try:
        parts = text.strip().split(":")
        if len(parts) != 2:
            return None
        return int(parts[0]), int(parts[1])
    except ValueError:
        return None


async def _recalculate_match(db, match_id: int) -> int:
    """Recalculate scores for all predictions on a match. Returns count updated."""
    match_row = await fetchone(db, "SELECT * FROM matches WHERE id = ?", (match_id,))
    if not match_row or match_row["status"] != "finished":
        return 0

    match = Match(
        score_home=match_row["score_home"],
        score_away=match_row["score_away"],
        went_to_extra_time=bool(match_row["went_to_extra_time"]),
        ot_pen_winner=match_row["ot_pen_winner"],
        ot_pen_method=match_row["ot_pen_method"],
    )
    predictions = await fetchall(
        db, "SELECT * FROM predictions WHERE match_id = ?", (match_id,)
    )
    for row in predictions:
        pred = Prediction(
            pred_home=row["pred_home"],
            pred_away=row["pred_away"],
            is_doubled=bool(row["is_doubled"]),
            playoff_team=row["playoff_team"],
            playoff_method=row["playoff_method"],
        )
        result = score_prediction(pred, match)
        await db.execute(
            """
            UPDATE predictions
            SET base_points = ?, base_final = ?, bonus_points = ?, total_points = ?
            WHERE id = ?
            """,
            (result["base_points"], result["base_final"],
             result["bonus_points"], result["total_points"], row["id"]),
        )
    await db.commit()
    return len(predictions)


async def _log_admin_action(db, match_id: int | None, action: str, payload: dict) -> None:
    payload["admin_tg"] = config.admin_telegram_id
    await db.execute(
        "INSERT INTO audit_log (user_id, match_id, action, payload, created_at) "
        "VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)",
        (config.admin_telegram_id, match_id or 0, action, json.dumps(payload, ensure_ascii=False)),
    )
    await db.commit()


def _pts_str(pts: int) -> str:
    a = abs(pts)
    if 11 <= a % 100 <= 19:
        form = "очков"
    elif a % 10 == 1:
        form = "очко"
    elif a % 10 in (2, 3, 4):
        form = "очка"
    else:
        form = "очков"
    return f"{pts:+d} {form}"


async def _publish_result(bot: Bot, match_id: int, db) -> None:
    """Post result to group chat (if configured) and notify each user individually."""
    match_row = await fetchone(db, "SELECT * FROM matches WHERE id = ?", (match_id,))
    if not match_row:
        return
    preds = await fetchall(
        db,
        """
        SELECT u.telegram_id, u.full_name,
               p.pred_home, p.pred_away, p.is_doubled,
               p.playoff_team, p.playoff_method,
               p.total_points, p.base_points
        FROM predictions p
        JOIN users u ON u.id = p.user_id
        WHERE p.match_id = ? AND p.total_points IS NOT NULL
        ORDER BY p.total_points DESC
        """,
        (match_id,),
    )
    ot_str = ""
    if match_row["went_to_extra_time"]:
        ot_str = f" ({match_row['ot_pen_winner']} через {match_row['ot_pen_method']})"
    header = (
        f"⚽ <b>{match_row['team_home']} {match_row['score_home']}:{match_row['score_away']}"
        f" {match_row['team_away']}</b>{ot_str}"
    )

    # Group chat summary
    if config.group_chat_id:
        lines = [header + "\n"]
        if preds:
            for p in preds:
                dbl = "×2 " if p["is_doubled"] else ""
                icon = "🎯" if p["base_points"] == 3 else ("✅" if p["base_points"] >= 1 else "❌")
                lines.append(
                    f"{icon} {p['full_name']}: {dbl}{p['pred_home']}:{p['pred_away']}"
                    f" → {p['total_points']:+d} очк."
                )
        else:
            lines.append("Прогнозов не было.")
        try:
            await bot.send_message(config.group_chat_id, "\n".join(lines))
        except Exception:
            pass

    # Individual notifications
    for p in preds:
        try:
            dbl = "×2 " if p["is_doubled"] else ""
            icon = "🎯" if p["base_points"] == 3 else ("✅" if p["base_points"] >= 1 else "❌")
            pred_str = f"{dbl}{p['pred_home']}:{p['pred_away']}"
            if p["playoff_team"]:
                m_str = f" {p['playoff_method']}" if p["playoff_method"] else ""
                pred_str += f", {p['playoff_team']}{m_str}"
            await bot.send_message(
                p["telegram_id"],
                f"{header}\n\nТвой прогноз: {pred_str}\n{icon} {_pts_str(p['total_points'])}",
            )
        except Exception:
            pass

    # Day summary (sends to all if this was the last match of the day)
    from bot.services.scheduler import _maybe_send_day_summary
    match_date = str(match_row["kickoff_msk"])[:10]
    await _maybe_send_day_summary(bot, match_date)


# ---------- /result — multi-step flow ----------

@router.message(Command("result"))
async def cmd_result(message: Message) -> None:
    args = (message.text or "").split(maxsplit=2)
    if len(args) >= 3:
        # Legacy: /result <match_id> <score>
        try:
            match_id = int(args[1])
        except ValueError:
            await message.answer("match_id должен быть числом.")
            return
        parsed = _parse_score(args[2])
        if not parsed:
            await message.answer("Неверный формат счёта. Пример: 2:1")
            return
        home, away = parsed
        async with get_db() as db:
            match = await fetchone(
                db, "SELECT id, team_home, team_away FROM matches WHERE id = ?", (match_id,)
            )
            if not match:
                await message.answer(f"Матч #{match_id} не найден.")
                return
            await db.execute(
                "UPDATE matches SET score_home=?, score_away=?, status='finished' WHERE id=?",
                (home, away, match_id),
            )
            await db.commit()
            count = await _recalculate_match(db, match_id)
            await _log_admin_action(db, match_id, "admin:result", {"score": f"{home}:{away}"})
            await _publish_result(message.bot, match_id, db)
        await message.answer(
            f"✅ Матч #{match_id} ({match['team_home']} — {match['team_away']}): "
            f"{home}:{away}\nПересчитано: {count}"
        )
        return

    # Interactive flow: show match list
    async with get_db() as db:
        rows = await fetchall(
            db,
            """SELECT id, team_home, team_away, kickoff_msk
               FROM matches WHERE status = 'scheduled'
               ORDER BY kickoff_msk LIMIT 20""",
        )
    if not rows:
        await message.answer("Нет матчей ожидающих результата.")
        return
    await message.answer("Выбери матч:", reply_markup=admin_match_list_kb(list(rows)))


@router.callback_query(F.data.startswith("arm:"))
async def cb_admin_match_select(callback: CallbackQuery) -> None:
    match_id = int(callback.data.split(":")[1])
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
    if not match:
        await callback.answer("Матч не найден.", show_alert=True)
        return
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} — {match['team_away']}</b>\n"
        f"Введи счёт {match['team_home']}:",
        reply_markup=admin_home_score_kb(match_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("ahd:"))
async def cb_admin_home_digit(callback: CallbackQuery) -> None:
    _, match_id_str, digit_str = callback.data.split(":")
    match_id, digit = int(match_id_str), int(digit_str)
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} — {match['team_away']}</b>\n"
        f"{match['team_home']}: {digit}\n"
        f"Введи счёт {match['team_away']}:",
        reply_markup=admin_away_score_kb(match_id, digit),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("aad:"))
async def cb_admin_away_digit(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    match_id, home_s, away_s = int(parts[1]), int(parts[2]), int(parts[3])
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away, stage FROM matches WHERE id = ?", (match_id,)
        )
    # If playoff match and draw → ask for OT/PEN winner
    if match["stage"] == "playoff" and home_s == away_s:
        await callback.message.edit_text(
            f"Матч: <b>{match['team_home']} — {match['team_away']}</b>\n"
            f"Счёт: {home_s}:{away_s} (ничья)\nКто прошёл дальше?",
            reply_markup=admin_playoff_team_kb(
                match_id, home_s, away_s, match["team_home"], match["team_away"]
            ),
        )
    else:
        # Go straight to confirmation
        await _show_admin_confirm(
            callback, match_id, home_s, away_s,
            match["team_home"], match["team_away"],
            winner="n", method="n",
        )
    await callback.answer()


@router.callback_query(F.data.startswith("apt:"))
async def cb_admin_playoff_team(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    match_id, home_s, away_s, winner = int(parts[1]), int(parts[2]), int(parts[3]), parts[4]
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
    team_label = match["team_home"] if winner == "h" else match["team_away"]
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} — {match['team_away']}</b>\n"
        f"Счёт: {home_s}:{away_s}  Победитель: {team_label}\nКак прошёл?",
        reply_markup=admin_playoff_method_kb(match_id, home_s, away_s, winner),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("apm:"))
async def cb_admin_playoff_method(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    match_id, home_s, away_s = int(parts[1]), int(parts[2]), int(parts[3])
    winner, method = parts[4], parts[5]  # h/a, OT/PEN/n
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
    await _show_admin_confirm(
        callback, match_id, home_s, away_s,
        match["team_home"], match["team_away"],
        winner=winner, method=method,
    )
    await callback.answer()


async def _show_admin_confirm(
    callback: CallbackQuery,
    match_id: int, home_s: int, away_s: int,
    team_home: str, team_away: str,
    winner: str, method: str,
) -> None:
    extra = ""
    if winner != "n":
        team_label = team_home if winner == "h" else team_away
        meth_label = "" if method == "n" else f" через {method}"
        extra = f"\nПобедитель: {team_label}{meth_label}"
    await callback.message.edit_text(
        f"Подтвердить результат?\n\n"
        f"Матч: <b>{team_home} — {team_away}</b>\n"
        f"Счёт: <b>{home_s}:{away_s}</b>{extra}",
        reply_markup=admin_confirm_kb(match_id, home_s, away_s, winner, method),
    )


@router.callback_query(F.data.startswith("acf:"))
async def cb_admin_confirm(callback: CallbackQuery) -> None:
    parts = callback.data.split(":")
    match_id, home_s, away_s = int(parts[1]), int(parts[2]), int(parts[3])
    winner, method = parts[4], parts[5]

    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
        if not match:
            await callback.answer("Матч не найден.", show_alert=True)
            return

        went_extra = winner != "n"
        ot_winner: str | None = None
        ot_method: str | None = None
        if went_extra:
            ot_winner = match["team_home"] if winner == "h" else match["team_away"]
            ot_method = None if method == "n" else method

        await db.execute(
            """UPDATE matches SET score_home=?, score_away=?, status='finished',
               went_to_extra_time=?, ot_pen_winner=?, ot_pen_method=? WHERE id=?""",
            (home_s, away_s, went_extra, ot_winner, ot_method, match_id),
        )
        await db.commit()
        count = await _recalculate_match(db, match_id)
        await _log_admin_action(
            db, match_id, "admin:result",
            {"score": f"{home_s}:{away_s}", "winner": ot_winner, "method": ot_method},
        )
        await _publish_result(callback.bot, match_id, db)

    extra = ""
    if ot_winner:
        meth_label = f" ({ot_method})" if ot_method else ""
        extra = f"\n🏆 Победитель: {ot_winner}{meth_label}"
    await callback.message.edit_text(
        f"✅ Результат сохранён\n"
        f"{match['team_home']} <b>{home_s}:{away_s}</b> {match['team_away']}{extra}\n"
        f"Пересчитано прогнозов: {count}"
    )
    await callback.answer()


@router.callback_query(F.data == "acancel")
async def cb_admin_cancel(callback: CallbackQuery) -> None:
    await callback.message.edit_text("Отменено.")
    await callback.answer()


# ---------- /playoff_result — interactive match list ----------

@router.message(Command("playoff_result"))
async def cmd_playoff_result(message: Message) -> None:
    async with get_db() as db:
        rows = await fetchall(
            db,
            """SELECT id, team_home, team_away, score_home, score_away, kickoff_msk
               FROM matches
               WHERE status = 'finished' AND went_to_extra_time = FALSE
                 AND team_home != '__adjustment__'
               ORDER BY kickoff_msk DESC LIMIT 20""",
        )
    if not rows:
        await message.answer("Нет завершённых матчей без записи об ОТ/ПЕН.")
        return
    await message.answer(
        "Выбери матч, в котором была ничья и нужно записать победителя по ОТ/ПЕН:",
        reply_markup=playoff_match_list_kb(list(rows)),
    )


@router.callback_query(F.data.startswith("pfm:"))
async def cb_playoff_match_select(callback: CallbackQuery) -> None:
    match_id = int(callback.data.split(":")[1])
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away, score_home, score_away FROM matches WHERE id = ?", (match_id,)
        )
    if not match:
        await callback.answer("Матч не найден.", show_alert=True)
        return
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} {match['score_home']}:{match['score_away']} {match['team_away']}</b>\n"
        f"Кто прошёл дальше?",
        reply_markup=playoff_pick_team_kb(match_id, match["team_home"], match["team_away"]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("ppt:"))
async def cb_playoff_pick_team(callback: CallbackQuery) -> None:
    _, match_id_str, winner = callback.data.split(":")
    match_id = int(match_id_str)
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away, score_home, score_away FROM matches WHERE id = ?", (match_id,)
        )
    team_label = match["team_home"] if winner == "h" else match["team_away"]
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} {match['score_home']}:{match['score_away']} {match['team_away']}</b>\n"
        f"Победитель: {team_label}\nКак прошёл?",
        reply_markup=playoff_pick_method_kb(match_id, winner),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("ppm:"))
async def cb_playoff_pick_method(callback: CallbackQuery) -> None:
    _, match_id_str, winner, method = callback.data.split(":")
    match_id = int(match_id_str)
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away, score_home, score_away FROM matches WHERE id = ?", (match_id,)
        )
    team_label = match["team_home"] if winner == "h" else match["team_away"]
    method_label = "доп. время (ОТ)" if method == "OT" else "пенальти (ПЕН)"
    await callback.message.edit_text(
        f"Подтвердить?\n\n"
        f"Матч: <b>{match['team_home']} {match['score_home']}:{match['score_away']} {match['team_away']}</b>\n"
        f"Победитель: <b>{team_label}</b> через {method_label}",
        reply_markup=playoff_confirm_kb(match_id, winner, method),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("ppc:"))
async def cb_playoff_confirm(callback: CallbackQuery) -> None:
    _, match_id_str, winner, method = callback.data.split(":")
    match_id = int(match_id_str)
    async with get_db() as db:
        match = await fetchone(
            db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,)
        )
        if not match:
            await callback.answer("Матч не найден.", show_alert=True)
            return
        team_name = match["team_home"] if winner == "h" else match["team_away"]
        await db.execute(
            "UPDATE matches SET went_to_extra_time=TRUE, ot_pen_winner=?, ot_pen_method=? WHERE id=?",
            (team_name, method, match_id),
        )
        await db.commit()
        count = await _recalculate_match(db, match_id)
        await _log_admin_action(
            db, match_id, "admin:playoff_result", {"winner": team_name, "method": method}
        )
        await _publish_result(callback.bot, match_id, db)
    method_label = "ОТ" if method == "OT" else "ПЕН"
    await callback.message.edit_text(
        f"✅ Записано: {team_name} через {method_label}\nПересчитано прогнозов: {count}"
    )
    await callback.answer()


# ---------- /recalc — with match list and confirmation ----------

@router.message(Command("recalc"))
async def cmd_recalc(message: Message) -> None:
    async with get_db() as db:
        rows = await fetchall(
            db,
            """SELECT id, team_home, team_away, score_home, score_away, kickoff_msk
               FROM matches WHERE status = 'finished' AND team_home != '__adjustment__'
               ORDER BY kickoff_msk DESC LIMIT 20""",
        )
    if not rows:
        await message.answer("Нет завершённых матчей.")
        return
    await message.answer("Выбери матч для пересчёта:", reply_markup=recalc_match_list_kb(list(rows)))


@router.callback_query(F.data.startswith("rpm:"))
async def cb_recalc_match_select(callback: CallbackQuery) -> None:
    match_id = int(callback.data.split(":")[1])
    async with get_db() as db:
        match = await fetchone(
            db,
            "SELECT team_home, team_away, score_home, score_away, status FROM matches WHERE id = ?",
            (match_id,),
        )
        pred_count = await fetchone(
            db, "SELECT COUNT(*) AS cnt FROM predictions WHERE match_id = ?", (match_id,)
        )
    if not match or match["status"] != "finished":
        await callback.answer("Матч не найден или не завершён.", show_alert=True)
        return
    count = pred_count["cnt"] if pred_count else 0
    await callback.message.edit_text(
        f"Пересчитать матч?\n\n"
        f"{match['team_home']} <b>{match['score_home']}:{match['score_away']}</b> {match['team_away']}\n"
        f"Прогнозов: {count}",
        reply_markup=recalc_confirm_kb(match_id),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("rcf:"))
async def cb_recalc_confirm(callback: CallbackQuery) -> None:
    match_id = int(callback.data.split(":")[1])
    async with get_db() as db:
        count = await _recalculate_match(db, match_id)
        await _log_admin_action(db, match_id, "admin:recalc", {})
    if count == 0:
        await callback.message.edit_text(f"Матч #{match_id} не найден или не завершён.")
    else:
        await callback.message.edit_text(f"✅ Пересчитано прогнозов: {count}")
    await callback.answer()


# ---------- /add_match + CSV upload ----------

@router.message(Command("add_match"))
async def cmd_add_match(message: Message) -> None:
    await message.answer(
        "Отправь CSV-файл с матчами.\n"
        "Колонки: match_date, kickoff_msk, stage, group_name, team_home, team_away"
    )


@router.message(F.document)
async def handle_csv_upload(message: Message) -> None:
    doc: Document = message.document
    if not doc.file_name.endswith(".csv"):
        return
    file = await message.bot.get_file(doc.file_id)
    data = await message.bot.download_file(file.file_path)
    text = data.read().decode("utf-8")
    reader = csv.DictReader(io.StringIO(text))
    required = {"match_date", "kickoff_msk", "stage", "team_home", "team_away"}
    if not required.issubset(set(reader.fieldnames or [])):
        await message.answer(f"CSV должен содержать колонки: {', '.join(sorted(required))}")
        return
    inserted = 0
    async with get_db() as db:
        for row in reader:
            await db.execute(
                """INSERT INTO matches (match_date, kickoff_msk, stage, group_name, team_home, team_away)
                   VALUES (?, ?, ?, ?, ?, ?)""",
                (row["match_date"], row["kickoff_msk"], row["stage"],
                 row.get("group_name") or None, row["team_home"], row["team_away"]),
            )
            inserted += 1
        await db.commit()
        await _log_admin_action(db, None, "admin:add_match", {"count": inserted})
    await message.answer(f"✅ Загружено матчей: {inserted}")


# ---------- /setup_api ----------

@router.message(Command("check_now"))
async def cmd_check_now(message: Message) -> None:
    """Trigger the daily API check immediately (for testing)."""
    from bot.services.scheduler import _daily_api_check
    await message.answer("⏳ Запускаю проверку API...")
    await _daily_api_check(message.bot)
    await message.answer("✅ Проверка завершена. Если есть новые результаты — см. выше.")


@router.message(Command("setup_api"))
async def cmd_setup_api(message: Message) -> None:
    """Register or re-authenticate with worldcup26.ir API and save token."""
    if not config.worldcup_api_email or not config.worldcup_api_password:
        await message.answer(
            "❌ WORLDCUP_API_EMAIL и WORLDCUP_API_PASSWORD не заданы в .env"
        )
        return
    try:
        await message.answer("⏳ Подключаюсь к API...")
        await setup_api()
        await message.answer("✅ Токен получен и сохранён. API готов к работе.")
    except Exception as exc:
        await message.answer(f"❌ Ошибка: {exc}")


# ---------- API auto-result callbacks ----------

@router.callback_query(F.data.startswith("apic:"))
async def cb_api_confirm(callback: CallbackQuery) -> None:
    """Confirm API result for a non-draw (or group) match — save score and recalculate."""
    parts = callback.data.split(":")
    match_id, home_s, away_s = int(parts[1]), int(parts[2]), int(parts[3])
    async with get_db() as db:
        match = await fetchone(db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,))
        if not match:
            await callback.answer("Матч не найден.", show_alert=True)
            return
        await db.execute(
            "UPDATE matches SET score_home=?, score_away=?, status='finished' WHERE id=?",
            (home_s, away_s, match_id),
        )
        await db.commit()
        count = await _recalculate_match(db, match_id)
        await _log_admin_action(
            db, match_id, "admin:result",
            {"score": f"{home_s}:{away_s}", "source": "api"},
        )
        await _publish_result(callback.bot, match_id, db)
    await callback.message.edit_text(
        f"✅ Результат сохранён (из API)\n"
        f"{match['team_home']} <b>{home_s}:{away_s}</b> {match['team_away']}\n"
        f"Пересчитано прогнозов: {count}"
    )
    await callback.answer()


@router.callback_query(F.data.startswith("apipo:"))
async def cb_api_playoff_open(callback: CallbackQuery) -> None:
    """
    API confirmed a playoff draw — save score, then open playoff winner selection.
    Reuses existing ppt: → ppm: → ppc: flow.
    """
    parts = callback.data.split(":")
    match_id, home_s, away_s = int(parts[1]), int(parts[2]), int(parts[3])
    async with get_db() as db:
        match = await fetchone(db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,))
        if not match:
            await callback.answer("Матч не найден.", show_alert=True)
            return
        await db.execute(
            "UPDATE matches SET score_home=?, score_away=?, status='finished' WHERE id=?",
            (home_s, away_s, match_id),
        )
        await db.commit()
        await _log_admin_action(
            db, match_id, "admin:result",
            {"score": f"{home_s}:{away_s}", "source": "api", "note": "draw_playoff"},
        )
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} {home_s}:{away_s} {match['team_away']}</b>\n"
        f"Счёт сохранён. Кто прошёл дальше?",
        reply_markup=playoff_pick_team_kb(match_id, match["team_home"], match["team_away"]),
    )
    await callback.answer()


@router.callback_query(F.data.startswith("apim:"))
async def cb_api_manual(callback: CallbackQuery) -> None:
    """Open manual score entry flow (reuses existing ahd: → aad: → ... chain)."""
    match_id = int(callback.data.split(":")[1])
    async with get_db() as db:
        match = await fetchone(db, "SELECT team_home, team_away FROM matches WHERE id = ?", (match_id,))
    if not match:
        await callback.answer("Матч не найден.", show_alert=True)
        return
    await callback.message.edit_text(
        f"Матч: <b>{match['team_home']} — {match['team_away']}</b>\n"
        f"Введи счёт {match['team_home']}:",
        reply_markup=admin_home_score_kb(match_id),
    )
    await callback.answer()


# ---------- API new-round callback ----------

@router.callback_query(F.data.startswith("apinr:"))
async def cb_api_new_round(callback: CallbackQuery) -> None:
    """Add all games of a new playoff round from API to our DB."""
    parts = callback.data.split(":", 2)
    round_type = parts[1]
    game_ids_str = parts[2] if len(parts) > 2 else ""
    game_ids = [gid.strip() for gid in game_ids_str.split(",") if gid.strip()]

    if not game_ids:
        await callback.answer("Нет матчей для добавления.", show_alert=True)
        return

    # Answer Telegram immediately — API fetch can take 10+ seconds
    await callback.answer()
    await callback.message.edit_text("⏳ Загружаю данные из API...")

    try:
        all_games = await fetch_games()
    except Exception as exc:
        await callback.message.edit_text(f"❌ Ошибка API: {exc}")
        return

    games_by_id = {str(g["id"]): g for g in all_games}
    inserted = 0
    async with get_db() as db:
        for gid in game_ids:
            g = games_by_id.get(gid)
            if not g or not g.get("home_team_name_en"):
                continue
            team_home = en_to_ru(g["home_team_name_en"])
            team_away = en_to_ru(g["away_team_name_en"])
            kickoff = local_to_msk(g["local_date"], g.get("stadium_id", ""))
            match_date = kickoff[:10]
            await db.execute(
                """INSERT INTO matches
                   (match_date, kickoff_msk, stage, team_home, team_away, api_game_id)
                   VALUES (?, ?, 'playoff', ?, ?, ?)""",
                (match_date, kickoff, team_home, team_away, gid),
            )
            inserted += 1
        await db.commit()
        await _log_admin_action(
            db, None, "admin:add_match",
            {"count": inserted, "round": round_type, "source": "api"},
        )

    round_label = {
        "r32": "1/16 финала", "r16": "1/8 финала",
        "qf": "Четвертьфинал", "sf": "Полуфинал", "final": "Финал",
    }.get(round_type, round_type.upper())
    await callback.message.edit_text(
        f"✅ Добавлено {inserted} матчей ({round_label}). Время конвертировано в МСК."
    )

    if inserted > 0:
        async with get_db() as db:
            all_users = await fetchall(db, "SELECT telegram_id FROM users")
        for u in all_users:
            try:
                await callback.bot.send_message(
                    u["telegram_id"],
                    f"🆕 <b>{round_label}</b> — открыты прогнозы!\n"
                    f"Добавлено матчей: {inserted}. Успей поставить до старта → /matches",
                )
            except Exception:
                pass


# ---------- /check_predictions ----------

@router.message(Command("check_predictions"))
async def cmd_check_predictions(message: Message) -> None:
    async with get_db() as db:
        total_users = (await fetchone(db, "SELECT COUNT(*) AS cnt FROM users"))["cnt"]
        matches = await fetchall(
            db,
            """SELECT m.id, m.team_home, m.team_away, m.kickoff_msk,
                      COUNT(p.id) AS pred_count
               FROM matches m
               LEFT JOIN predictions p ON p.match_id = m.id
               WHERE m.status = 'scheduled' AND m.team_home != '__adjustment__'
               GROUP BY m.id
               ORDER BY m.kickoff_msk
               LIMIT 15""",
        )
        if not matches:
            await message.answer("Нет предстоящих матчей.")
            return

        lines = [f"📋 <b>Прогнозы на предстоящие матчи</b> (всего участников: {total_users}):\n"]
        for m in matches:
            pc = m["pred_count"]
            missing = total_users - pc
            bar = "✅" * pc + "⬜" * missing
            time_str = str(m["kickoff_msk"])[5:16].replace("-", ".")
            lines.append(
                f"<b>{m['team_home']} — {m['team_away']}</b>  {time_str}\n"
                f"{bar}  {pc}/{total_users}"
            )
            if missing > 0:
                # Show names of users who haven't predicted
                no_pred = await fetchall(
                    db,
                    """SELECT u.full_name FROM users u
                       WHERE NOT EXISTS (
                           SELECT 1 FROM predictions p
                           WHERE p.user_id = u.id AND p.match_id = ?
                       )""",
                    (m["id"],),
                )
                names = ", ".join(r["full_name"] for r in no_pred)
                lines.append(f"<i>Не поставили: {names}</i>")
            lines.append("")

    await message.answer("\n".join(lines).rstrip())


# ---------- /export ----------

@router.message(Command("export"))
async def cmd_export(message: Message) -> None:
    """Generate and send an Excel table matching the original Битва прогнозов format."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
    except ImportError:
        await message.answer("⚠️ Нет openpyxl: pip install openpyxl")
        return

    from io import BytesIO
    from datetime import date as _date
    from collections import defaultdict
    from aiogram.types import BufferedInputFile

    await message.answer("⏳ Формирую таблицу...")

    async with get_db() as db:
        users_raw = await fetchall(db, "SELECT id, full_name, doublings_left FROM users")
        pts_rows = await fetchall(
            db,
            "SELECT user_id, COALESCE(SUM(total_points),0) AS pts "
            "FROM predictions WHERE total_points IS NOT NULL GROUP BY user_id",
        )
        pts_map: dict[int, int] = {r["user_id"]: r["pts"] for r in pts_rows}
        exact_rows = await fetchall(
            db,
            """SELECT p.user_id, COUNT(*) AS cnt FROM predictions p
               JOIN matches m ON m.id = p.match_id
               WHERE p.base_points = 3 AND m.team_home != '__adjustment__'
                 AND p.total_points IS NOT NULL
               GROUP BY p.user_id""",
        )
        exact_map: dict[int, int] = {r["user_id"]: r["cnt"] for r in exact_rows}
        users = sorted(
            users_raw,
            key=lambda u: (-pts_map.get(u["id"], 0), -exact_map.get(u["id"], 0)),
        )
        user_ids = [u["id"] for u in users]

        matches = await fetchall(
            db,
            "SELECT * FROM matches WHERE team_home != '__adjustment__' ORDER BY kickoff_msk",
        )
        preds_raw = await fetchall(
            db,
            """SELECT p.user_id, p.match_id, p.pred_home, p.pred_away,
                      p.is_doubled, p.total_points, p.base_points
               FROM predictions p JOIN matches m ON m.id = p.match_id
               WHERE m.team_home != '__adjustment__'""",
        )
        group_rows = await fetchall(
            db,
            """SELECT p.user_id, COALESCE(SUM(p.total_points),0) AS pts
               FROM predictions p JOIN matches m ON m.id = p.match_id
               WHERE m.stage = 'group' AND p.total_points IS NOT NULL
               GROUP BY p.user_id""",
        )
        group_pts_map: dict[int, int] = {r["user_id"]: r["pts"] for r in group_rows}

    pred_map: dict[tuple[int, int], dict] = {
        (p["user_id"], p["match_id"]): p for p in preds_raw
    }
    match_date_map = {m["id"]: str(m["kickoff_msk"])[:10] for m in matches}
    day_pts: dict[str, dict[int, int]] = defaultdict(lambda: defaultdict(int))
    for p in preds_raw:
        if p["total_points"] is not None:
            dk = match_date_map.get(p["match_id"], "")
            if dk:
                day_pts[dk][p["user_id"]] += p["total_points"]

    # ── Styles ─────────────────────────────────────────────────────────────
    N = len(users)
    _thin = Side(style="thin")
    _border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
    _orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    _day_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    _group_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    _grand_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    _bold = Font(bold=True)

    # ── Build workbook ──────────────────────────────────────────────────────
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "ЧМ-2026"
    today_str = _date.today().strftime("%d.%m")

    def _add_row(values: list, fill=None, font=None, orange_cols: list[int] | None = None) -> None:
        """Append a row; apply border, optional fill/font, and orange on specific 1-based cols."""
        ws.append(values)
        ri = ws.max_row
        for ci in range(1, len(values) + 1):
            cell = ws.cell(row=ri, column=ci)
            cell.border = _border
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
        if orange_cols:
            for ci in orange_cols:
                ws.cell(row=ri, column=ci).fill = _orange

    def _total_row(label: str, pts_by_uid: dict[int, int], fill=None, font=None) -> None:
        vals: list = [label, "", "", ""]
        s = 0
        for uid in user_ids:
            v = pts_by_uid.get(uid, 0)
            vals.append(v if v != 0 else "")
            s += v
        vals.append(s if s != 0 else "")
        _add_row(vals, fill=fill, font=font)

    # ── Header rows (all have N+5 columns: A-D + N players + Σ) ───────────
    _add_row(["", "", "Количество точных счётов", ""]
             + [exact_map.get(u["id"], 0) for u in users] + [""],
             font=_bold)
    _add_row(["", "", f"Осталось удвоений (обн. {today_str})", ""]
             + [u["doublings_left"] for u in users] + [""],
             font=_bold)
    _add_row(["Дата", "Время (Мск)", "Матч", "Итог матча"]
             + [u["full_name"] for u in users] + ["Σ"],
             font=_bold)

    ws.freeze_panes = "A4"

    # ── Match rows ─────────────────────────────────────────────────────────
    prev_date: str | None = None
    group_done = False

    for m in matches:
        m_date = str(m["kickoff_msk"])[:10]
        m_time = str(m["kickoff_msk"])[11:16]

        if not group_done and m["stage"] == "playoff":
            group_done = True
            _total_row("Итог группового этапа",
                       {uid: group_pts_map.get(uid, 0) for uid in user_ids},
                       fill=_group_fill, font=_bold)

        if prev_date and prev_date != m_date:
            _total_row("Итог игрового дня", day_pts[prev_date], fill=_day_fill)

        date_fmt = m_date[8:] + "." + m_date[5:7] + "." + m_date[:4]
        h, mn = m_time.split(":")
        time_fmt = f"{int(h)}:{mn}"
        match_name = f"{m['team_home']} – {m['team_away']}"
        if m["status"] == "finished" and m["score_home"] is not None:
            result = f"{m['score_home']}:{m['score_away']}"
            if m["went_to_extra_time"] and m["ot_pen_winner"]:
                result += f" ({m['ot_pen_winner']} {m['ot_pen_method']})"
        else:
            result = ""

        row: list = [date_fmt, time_fmt, match_name, result]
        orange_cols: list[int] = []
        for i, uid in enumerate(user_ids):
            pred = pred_map.get((uid, m["id"]))
            row.append(
                f"{pred['pred_home']}:{pred['pred_away']}"
                if pred and pred["pred_home"] is not None else ""
            )
            if pred and pred["is_doubled"]:
                orange_cols.append(5 + i)  # 1-based column index
        row.append("")  # Σ column empty for match rows
        _add_row(row, orange_cols=orange_cols or None)
        prev_date = m_date

    if prev_date:
        _total_row("Итог игрового дня", day_pts[prev_date], fill=_day_fill)

    if not group_done:
        _total_row("Итог группового этапа",
                   {uid: group_pts_map.get(uid, 0) for uid in user_ids},
                   fill=_group_fill, font=_bold)

    _total_row("Итог битвы", pts_map, fill=_grand_fill, font=_bold)

    # ── Column widths ───────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 10
    ws.column_dimensions["C"].width = 34
    ws.column_dimensions["D"].width = 22
    for i in range(N):
        ws.column_dimensions[openpyxl.utils.get_column_letter(5 + i)].width = 12
    ws.column_dimensions[openpyxl.utils.get_column_letter(5 + N)].width = 6

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    today_full = _date.today().strftime("%d.%m.%Y")
    await message.answer_document(
        BufferedInputFile(buf.read(), filename=f"Битва_прогнозов_ЧМ2026_{today_full}.xlsx"),
        caption=f"📊 Таблица прогнозов ЧМ-2026 на {today_full}",
    )


# ---------- /admin_log ----------

@router.message(Command("admin_log"))
async def cmd_admin_log(message: Message) -> None:
    async with get_db() as db:
        rows = await fetchall(
            db,
            """SELECT a.action, a.match_id, a.payload, a.created_at,
                      m.team_home, m.team_away
               FROM audit_log a
               LEFT JOIN matches m ON m.id = a.match_id
               WHERE a.action LIKE 'admin:%'
               ORDER BY a.id DESC LIMIT 10""",
        )
    if not rows:
        await message.answer("Лог действий пуст.")
        return
    lines = ["📋 <b>Последние действия:</b>\n"]
    for r in rows:
        ts = str(r["created_at"])[:16]
        try:
            payload = json.loads(r["payload"] or "{}")
        except Exception:
            payload = {}
        action = r["action"].replace("admin:", "")
        match_label = (
            f" [{r['team_home']} — {r['team_away']}]"
            if r["team_home"] else ""
        )
        detail = ""
        if action == "result":
            detail = f" → {payload.get('score', '')}"
        elif action == "playoff_result":
            detail = f" → {payload.get('winner', '')} ({payload.get('method', '')})"
        elif action == "add_match":
            detail = f" ({payload.get('count', '')} матчей)"
        lines.append(f"[{ts}] {action}{match_label}{detail}")
    await message.answer("\n".join(lines))
